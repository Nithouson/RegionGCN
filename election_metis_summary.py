import openpyxl
import math
import numpy as np


def normalized_mutual_info(reg, rreg):
    n = sum([len(r) for r in reg])
    assert n == sum([len(r) for r in rreg])
    h = -sum([(len(r) / n) * math.log2(len(r) / n) for r in reg])
    hr = -sum([(len(r) / n) * math.log2(len(r) / n) for r in rreg])
    reg_set = [set(r) for r in reg]
    rreg_set = [set(r) for r in rreg]
    mi = 0
    for r in reg_set:
        for rr in rreg_set:
            n_int = len(r.intersection(rr))
            if n_int == 0:
                continue
            mi += (n_int / n) * math.log2(n * n_int / (len(r) * len(rr)))
    nmi = 2 * mi / (h + hr)
    return nmi


n_units = 3108
units = np.arange(n_units).astype(int)

# Load Reference Regions
reg_wb = openpyxl.load_workbook("./log/Election_sys4/reggcn_zones_4.xlsx")
reg_ws = reg_wb['regions']
assert reg_ws.max_row == n_units + 1
ref_regions = []
for c in range(3, 13):
    ref_rlabel = np.asarray([reg_ws.cell(r, c).value for r in range(2, n_units+2)], dtype=int)
    label_set = set(ref_rlabel)
    ref_reg = [units[ref_rlabel == r].tolist() for r in label_set]
    ref_regions.append(ref_reg)
reg_wb.close()

# Generate Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(1, 1, "ID")
for u in range(n_units):
    ws.cell(2 + u, 1, u)
col = 2
uflist = [1000, 3000, 5000]  #[500, 1000, 2000, 3000, 5000, 10000, 50000]
for uf in uflist:
    for reg in range(5, 55, 5):
        file = open(f"../graphcut/results/Election_sys4/part{reg}_u{uf}-reggcn4.txt")
        rlabel = np.asarray([-1 for u in range(n_units)],dtype=int)
        for l in range(n_units):
            args = file.readline().split(" ")
            u, r = int(args[0]), int(args[1])
            rlabel[u] = r
        ws.cell(1, col, f"u{uf}r{reg}")

        # Calculate ANMI
        label_set = set(rlabel)
        regions = [units[rlabel == r].tolist() for r in label_set]
        anmi = sum([normalized_mutual_info(regions, ref_reg) for ref_reg in ref_regions])\
               /len(ref_regions)
        print(uf, reg, anmi)

        # Write Data into Workbook
        for u in range(n_units):
            assert rlabel[u] >= 0
            ws.cell(2 + u, col, rlabel[u])
        col += 1
wb.save("../graphcut/results/reggcn4_utest.xlsx")


