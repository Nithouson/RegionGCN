import openpyxl
import copy
from libpysal import weights
from reggcn import weights_to_graph


def common_region(ra, rb):
    assert len(ra) == len(rb)
    c = 0
    for i in range(len(ra)):
        if ra[i] == rb[i]:
            c += 1
    return c


rook = weights.Rook.from_shapefile('cb_2016_cus_county_500k.shp', idVariable='GEOID')
links = copy.copy(rook.neighbors)
links['25001'] += ['25007', '25019']  # Barnstable County, MA
links['25007'] += ['25001', '25019']  # Dukes County, MA
links['25019'] += ['25001', '25007']  # Nantucket County, MA
links['53055'] += ['53057']  # San Juan Islands, WA
links['53057'] += ['53055']  # Skagit County, WA
links_id = {}
for unit in links.keys():
    uid = rook.id2i[unit]
    links_id[uid] = [rook.id2i[nunit] for nunit in links[unit]]
w = weights.W(links_id)
print(w.n_components, len(w.islands))  # connected
g = weights_to_graph(w)
nodes = g.number_of_nodes()
print(nodes)

reg_wb = openpyxl.load_workbook("zone_file.xlsx")
# a file recording multiple zoning results, each row for a county and column for a partition
reg_ws = reg_wb['regions']
assert reg_ws.max_row == nodes + 1
reg_dict = dict()
for r in range(2, nodes + 2):
    id = reg_ws.cell(r, 1).value
    region = [reg_ws.cell(r, c).value for c in range(3, 13)]
    # column 3,4,...,12 are region indexes
    reg_dict[id] = region

mf = open("metis_graph.txt","w")
mf.write(f"{g.number_of_nodes()} {g.number_of_edges()} 001\n")
for v in range(nodes):
    edges = dict()
    for vn in links_id[v]:
        edges[vn] = common_region(reg_dict[v], reg_dict[vn])
    mf.write(" ".join([f"{vn} {edges[vn]}" for vn in edges.keys()]))
    mf.write("\n")
mf.close()
